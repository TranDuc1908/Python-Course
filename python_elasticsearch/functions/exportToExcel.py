# encoding: utf-8
import openpyxl
from python_elasticsearch.coreFunction.coreHelper import CoreHelper

class exportToExcel(CoreHelper):
  def __init__(self):
    self.filename = "example1.xlsx"
    self.wb = openpyxl.Workbook()
    self.sheet = self.wb.active
  
  def getMaxRow(self):
    self.sheet['A1'] = "Tiêu đề"
    self.sheet['B1'] = "Danh mục"
    self.sheet['C1'] = "Mô tả"
    self.sheet['D1'] = "Nội dung"
    self.sheet['E1'] = "Tác giả"
    self.sheet['F1'] = "Thời gian đăng bài"
    self.sheet['G1'] = "Thời gian lấy về"
    self.wb.save(self.filename)

    self.wb = openpyxl.load_workbook(self.filename)
    self.sheet = self.wb.worksheets[0]
    return self.sheet.max_row

  def exportToExcel(self, _list):
    count = self.getMaxRow() + 1

    for item in _list:
      self.sheet.cell(row=count, column=1).hyperlink = item["_source"]["url"]
      self.sheet.cell(row=count, column=1).value = item["_source"]["title"]
      self.sheet.cell(row=count, column=1).style = "Hyperlink"

      self.sheet.cell(row=count, column=2).hyperlink = item["_source"]["category_url"]
      self.sheet.cell(row=count, column=2).value = item["_source"]["category_title"]
      self.sheet.cell(row=count, column=2).style = "Hyperlink"

      self.sheet['C'+str(count)] = item["_source"]["intro"]

      _str = ""
      for z in item["_source"]["content"]:
        _str += self.unicodeTrans(z)
      self.sheet['D'+str(count)] = _str

      self.sheet['E'+str(count)] = item["_source"]["author"]
      self.sheet['F'+str(count)] = item["_source"]["published_at"]
      self.sheet['G'+str(count)] = item["_source"]["created_at"]
      count += 1

    self.wb.save(self.filename)

